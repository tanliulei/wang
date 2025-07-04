import streamlit as st
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
import tempfile
import traceback
from io import BytesIO
from streamlit_extras.app_logo import add_logo
from streamlit_extras.metric_cards import style_metric_cards

st.set_page_config(
    page_title="PDF转Excel神器",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)
add_logo("https://img.icons8.com/ios-filled/100/000000/pdf.png", height=60)

st.markdown(
    """
    <style>
    .main {
        background: linear-gradient(135deg, #f8fafc 0%, #e0e7ef 100%);
    }
    .stButton>button {
        background: linear-gradient(90deg, #4f8cff 0%, #38b6ff 100%);
        color: white;
        font-weight: bold;
        border-radius: 8px;
        padding: 0.5em 2em;
        margin-top: 1em;
    }
    .stDownloadButton>button {
        background: #38b6ff;
        color: white;
        font-weight: bold;
        border-radius: 8px;
        padding: 0.5em 2em;
    }
    .stProgress>div>div>div>div {
        background-image: linear-gradient(90deg, #4f8cff, #38b6ff);
    }
    .stDataFrame {
        background: #fff;
        border-radius: 8px;
        box-shadow: 0 2px 8px #e0e7ef;
    }
    .stSidebar {
        background: #f0f4fa;
    }
    </style>
    """,
    unsafe_allow_html=True
)

def extract_pdf_to_dataframe(pdf_file):
    """
    Extract data from PDF file and convert to DataFrame
    """
    try:
        tables = []
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                # Try to extract table from the page
                page_tables = page.extract_tables()
                if page_tables:
                    for table in page_tables:
                        if table:  # Check if table is not empty
                            tables.extend(table)
                else:
                    # If no tables found, try to extract text and split by lines
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines:
                            if line.strip():
                                # Split by common delimiters
                                row = line.split('\t') if '\t' in line else line.split()
                                if len(row) > 1:
                                    tables.append(row)
        if not tables:
            st.error("无法从PDF中提取任何数据。请确保PDF包含表格或结构化数据。")
            return None
        # Convert to DataFrame
        df = pd.DataFrame(tables)
        # Clean up the DataFrame - remove empty rows and columns
        df = df.dropna(how='all').dropna(axis=1, how='all')
        # Reset index
        df = df.reset_index(drop=True)
        return df
    except Exception as e:
        st.error(f"PDF提取失败: {str(e)}")
        return None

def process_excel_data(df):
    """
    Process the DataFrame according to the specified requirements
    """
    try:
        # Make a copy to avoid modifying original data
        processed_df = df.copy()
        
        # Ensure we have enough columns (at least I column which is index 8)
        if len(processed_df.columns) < 9:
            # Add empty columns if needed
            for i in range(len(processed_df.columns), 9):
                processed_df[i] = ''
        
        # Delete column A (交易单号), H, I (indices 0, 7, 8)
        # After deletion, 交易时间 will become the new A column
        columns_to_drop = []
        if 0 < len(processed_df.columns):  # Column A (index 0) - 交易单号
            columns_to_drop.append(0)
        if 7 < len(processed_df.columns):  # Column H (index 7)
            columns_to_drop.append(7)
        if 8 < len(processed_df.columns):  # Column I (index 8)
            columns_to_drop.append(8)
        
        # Drop columns in reverse order to maintain indices
        for col_idx in sorted(columns_to_drop, reverse=True):
            processed_df = processed_df.drop(processed_df.columns[col_idx], axis=1)
        
        # Reset column names
        processed_df.columns = range(len(processed_df.columns))
        
        # Process columns: G (交易对方) and B (交易时间)
        # After deleting A (交易单号), B (交易时间) becomes index 0, G (交易对方) becomes index 5
        b_column_index = 0  # 交易时间 is now the first column (new A column)
        g_column_index = 5 if len(processed_df.columns) > 5 else len(processed_df.columns) - 1  # 交易对方
        
        if g_column_index < len(processed_df.columns) and b_column_index < len(processed_df.columns):
            # Sort by column G (交易对方) first, then by column B (交易时间) from newest to oldest
            # For time sorting, we'll try to convert to datetime if possible
            try:
                # Convert to datetime and sort first
                temp_datetime = pd.to_datetime(processed_df[b_column_index], errors='coerce')
                
                # Sort by G column (交易对方) first, then by datetime
                processed_df = processed_df.sort_values(
                    by=[g_column_index, b_column_index], 
                    ascending=[True, False],  # G列升序(相同交易对方聚集), B列降序(时间从近到远)
                    na_position='last'
                )
                
                # Format datetime to remove seconds - apply to all cells in the column
                def format_time(time_str):
                    try:
                        if pd.isna(time_str) or time_str == '':
                            return time_str
                        dt = pd.to_datetime(time_str, errors='coerce')
                        if pd.isna(dt):
                            return time_str
                        return dt.strftime('%Y-%m-%d %H:%M')
                    except:
                        return time_str
                
                processed_df[b_column_index] = processed_df[b_column_index].apply(format_time)
                
            except:
                # If datetime conversion fails, sort as strings
                processed_df = processed_df.sort_values(
                    by=[g_column_index, b_column_index], 
                    ascending=[True, False],  # G列升序(相同交易对方聚集), B列降序(时间从近到远)
                    na_position='last'
                )
            
            processed_df = processed_df.reset_index(drop=True)
        
        return processed_df
    
    except Exception as e:
        st.error(f"数据处理失败: {str(e)}")
        return None

def save_to_excel(df, save_path):
    """
    保存DataFrame到Excel文件，并对A列列宽设置为原始宽度的2倍，A列日期和时间用空格隔开。第5列连续两行及以上的固定数值且大于等于80，且这些行第3列内容含“支出”，则将这些区段的第1列、第3列、第5列、第6列字体标红。
    """
    try:
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            # 先处理A列内容：将日期和时间用空格隔开
            df_mod = df.copy()
            if df_mod.shape[1] >= 1:
                def date_time_space(val):
                    if isinstance(val, str):
                        # 先将逗号替换为空格
                        val = val.replace(',', ' ')
                        # 如果有T分隔符也替换为空格
                        val = val.replace('T', ' ')
                    return val
                df_mod.iloc[:, 0] = df_mod.iloc[:, 0].apply(date_time_space)
            df_mod.to_excel(writer, index=False, header=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, PatternFill, Font

            # 设置A列列宽为原始宽度2倍
            if worksheet.max_column >= 1:
                col_letter = get_column_letter(1)
                worksheet.column_dimensions[col_letter].width = 8.4 * 2

            # 设置第三列列宽为原始宽度0.8倍，字体居中
            if worksheet.max_column >= 3:
                col_letter = get_column_letter(3)
                worksheet.column_dimensions[col_letter].width = 6.7  # 8.4*0.8
                for row in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=3)
                    cell.alignment = Alignment(horizontal='center')

            # 新标红逻辑：第5列连续两行及以上的固定数值且>=80，且这些行第3列内容含“支出”，标红第1、3、5、6列
            amount_col = 5
            target_col = 6
            time_col = 1
            expense_col = 3
            red_font = Font(color='FFFF0000')
            data_amount = [worksheet.cell(row=i, column=amount_col).value for i in range(1, worksheet.max_row + 1)]
            data_expense = [worksheet.cell(row=i, column=expense_col).value for i in range(1, worksheet.max_row + 1)]
            n = len(data_amount)
            i = 0
            while i < n:
                val = data_amount[i]
                try:
                    num = float(str(val).replace(',', '').replace(' ', ''))
                except:
                    i += 1
                    continue
                # 判断是否支出项（第3列内容含“支出”）
                exp_val = str(data_expense[i]) if i < len(data_expense) else ''
                is_expense = ('支出' in exp_val)
                if num < 80 or not is_expense:
                    i += 1
                    continue
                # 检查后续有多少行相同且第3列为支出项
                j = i + 1
                while j < n:
                    try:
                        next_num = float(str(data_amount[j]).replace(',', '').replace(' ', ''))
                    except:
                        break
                    exp_val_j = str(data_expense[j]) if j < len(data_expense) else ''
                    is_expense_j = ('支出' in exp_val_j)
                    if next_num == num and is_expense_j:
                        j += 1
                    else:
                        break
                if j - i >= 2:
                    for k in range(i, j):
                        worksheet.cell(row=k+1, column=time_col).font = red_font
                        worksheet.cell(row=k+1, column=expense_col).font = red_font
                        worksheet.cell(row=k+1, column=amount_col).font = red_font
                        worksheet.cell(row=k+1, column=target_col).font = red_font
                    i = j
                else:
                    i += 1
            # 对第五列数值大于等于5000的金额和交易对方字体标橙（优先级低于红色）
            orange_font = Font(color='FFFF9900')
            for row in range(1, worksheet.max_row + 1):
                cell_amount = worksheet.cell(row=row, column=amount_col)
                cell_target = worksheet.cell(row=row, column=target_col)
                try:
                    num = float(str(cell_amount.value).replace(',', '').replace(' ', ''))
                except:
                    continue
                is_red = cell_amount.font and cell_amount.font.color and cell_amount.font.color.rgb == 'FFFF0000'
                if num >= 5000 and not is_red:
                    cell_amount.font = orange_font
                    cell_target.font = orange_font
        return True
    except Exception as e:
        st.error(f"Excel文件保存失败: {str(e)}")
        return False

def extract_pdf_title_name(pdf_file):
    """
    从PDF首页提取“兹证明：”后面的姓名作为Excel文件名。
    """
    try:
        with pdfplumber.open(pdf_file) as pdf:
            first_page = pdf.pages[0]
            text = first_page.extract_text()
            if text:
                lines = text.split('\n')
                # 优先找“兹证明：”
                for line in lines:
                    if '兹证明：' in line:
                        # 取“兹证明：XXX”格式
                        parts = line.split('兹证明：', 1)
                        if len(parts) == 2:
                            name = parts[1].strip()
                            # 去除后续非姓名字符
                            name = name.split()[0] if name else '导出文件'
                            return name
                # 兜底：找“姓名”
                for line in lines:
                    if '姓名' in line:
                        parts = line.replace('：', ':').split(':')
                        if len(parts) == 2:
                            return parts[1].strip()
                        else:
                            return line.strip().replace('姓名', '').strip()
                # 否则取第一行
                return lines[0].strip()
    except Exception:
        pass
    return '导出文件'

def main():
    st.title("📄 PDF到Excel转换神器")
    st.markdown("---")
    style_metric_cards(background_color="#e0e7ef", border_left_color="#38b6ff")
    st.header("1. 上传PDF文件")
    uploaded_file = st.file_uploader(
        "选择PDF文件", 
        type="pdf",
        help="支持微信支付明细PDF文件，最大200MB"
    )
    if uploaded_file is not None:
        file_data = uploaded_file.getvalue()
        file_size_mb = len(file_data) / (1024 * 1024)
        st.info(f"文件名: {uploaded_file.name}")
        st.info(f"文件大小: {file_size_mb:.1f}MB")
        if file_size_mb > 200:
            st.error("文件过大，请选择小于200MB的文件")
        elif file_size_mb == 0:
            st.error("文件为空，请选择有效的PDF文件")
        elif not file_data.startswith(b'%PDF'):
            st.error("文件格式不正确，请选择有效的PDF文件")
        else:
            st.success("文件验证通过，自动开始分析...")
            import tempfile as temp_module
            with temp_module.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_file_path = tmp_file.name
            try:
                st.subheader("2. 分析进度")
                progress_bar = st.progress(0)
                status_text = st.empty()
                # 步骤1：提取PDF数据
                status_text.text("正在提取PDF数据...")
                progress_bar.progress(20)
                df = extract_pdf_to_dataframe(tmp_file_path)
                # 新增：提取PDF抬头姓名
                excel_title = extract_pdf_title_name(tmp_file_path)
                if df is not None:
                    progress_bar.progress(40)
                    status_text.text("PDF数据提取完成")
                    st.subheader("3. 数据预览")
                    st.dataframe(df.head(10))
                    st.info(f"提取到 {len(df)} 行数据，{len(df.columns)} 列")
                    # 步骤2：处理数据
                    progress_bar.progress(60)
                    status_text.text("正在处理Excel数据...")
                    processed_df = process_excel_data(df)
                    if processed_df is not None:
                        progress_bar.progress(80)
                        status_text.text("数据处理完成")
                        st.subheader("4. 处理后的数据预览")
                        st.dataframe(processed_df.head(10))
                        st.info(f"处理后有 {len(processed_df)} 行数据，{len(processed_df.columns)} 列")
                        # 步骤3：保存Excel
                        progress_bar.progress(90)
                        status_text.text("正在保存Excel文件...")
                        save_dir = os.path.expanduser("~/Downloads")
                        if not os.path.exists(save_dir):
                            save_dir = temp_module.gettempdir()
                        # 用抬头姓名作为文件名
                        excel_filename = f"{excel_title}.xlsx"
                        save_path = os.path.join(save_dir, excel_filename)
                        counter = 1
                        while os.path.exists(save_path):
                            excel_filename = f"{excel_title}_{counter}.xlsx"
                            save_path = os.path.join(save_dir, excel_filename)
                            counter += 1
                        if save_to_excel(processed_df, save_path):
                            progress_bar.progress(100)
                            status_text.text(f"处理完成！已自动保存到 {save_path}")
                            st.success(f"✅ 文件处理完成！已自动保存到 {save_path}")
                            with open(save_path, 'rb') as f:
                                excel_data = f.read()
                            st.download_button(
                                label="下载Excel文件",
                                data=excel_data,
                                file_name=excel_filename,
                                mime="application/vnd.ms-excel"
                            )
                else:
                    progress_bar.progress(0)
                    status_text.text("PDF数据提取失败")
            except Exception as e:
                st.error(f"处理过程中发生错误: {str(e)}")
                st.error("详细错误信息:")
                st.code(traceback.format_exc())
            finally:
                try:
                    os.unlink(tmp_file_path)
                except:
                    pass
    else:
        st.info("👆 请上传PDF文件开始处理")
        st.subheader("使用说明")
        st.markdown("""
        1. **上传PDF文件**: 点击上方的文件上传按钮选择PDF文件
        2. **极速转换**: 应用会自动将PDF转换为Excel格式
        3. **数据处理**: 执行自动化操作
        4. **保存文件**: 处理后的文件将自动保存到下载目录
        5. **下载选项**: 提供文件下载功能
        """)
        st.subheader("注意事项")
        st.warning("""
        - 请确保PDF文件包含表格或结构化数据
        - 处理大文件可能需要较长时间
        - 确保有足够的磁盘空间保存处理后的文件
        - 如果指定保存路径不存在，文件将保存到应用程序目录
        """)

if __name__ == "__main__":
    main()
